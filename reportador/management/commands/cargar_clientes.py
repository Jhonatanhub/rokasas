import os
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
# Cambia 'tu_app' por el nombre real de tu aplicación de Django (ej. reportador)
from reportador.models import Cliente, ClienteCorreo  

class Command(BaseCommand):
    help = 'Carga masiva de clientes y correos desde el archivo original de Excel (.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta absoluta o relativa al archivo .xlsx')

    def handle(self, *args, **options):
        excel_path = options['excel_file']

        if not os.path.exists(excel_path):
            self.stdout.write(self.style.ERROR(f"El archivo en la ruta '{excel_path}' no existe."))
            return

        self.stdout.write(self.style.SUCCESS(f"Iniciando lectura del archivo Excel: {excel_path}"))

        try:
            # Cargamos el libro de trabajo y seleccionamos la primera hoja activa
            wb = load_workbook(filename=excel_path, data_only=True)
            sheet = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error al abrir el archivo Excel: {str(e)}"))
            return

        headers = {}
        header_row_index = None

        # 1. Identificar la fila de encabezados reales
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Limpiamos y normalizamos los valores de la fila actual
            clean_cells = [str(cell).strip().lower() if cell is not None else '' for cell in row]
            
            if 'nit' in clean_cells and 'detalle terceros' in clean_cells:
                header_row_index = row_idx
                # Mapeamos el nombre de la columna con su índice (posición)
                headers = {str(cell).strip(): idx for idx, cell in enumerate(row) if cell is not None}
                break

        if not header_row_index:
            self.stdout.write(self.style.ERROR("No se encontró una fila de encabezados válida que contenga 'Nit' y 'Detalle Terceros'."))
            return

        self.stdout.write(self.style.SUCCESS(f"Encabezados encontrados en la fila {header_row_index}."))

        clientes_creados = 0
        correos_registrados = 0

        # 2. Iterar las filas a partir de la siguiente línea de los encabezados
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
            
            # Obtener el Nit y el Nombre usando el mapeo de encabezados
            nit_val = row[headers.get('Nit')] if headers.get('Nit') is not None else None
            nombre_val = row[headers.get('Detalle Terceros')] if headers.get('Detalle Terceros') is not None else None

            # Si el NIT o el nombre vienen vacíos, saltamos la fila (control de filas vacías al final)
            if not nit_val or not nombre_val:
                continue

            nit = str(nit_val).strip()
            nombre = str(nombre_val).strip()

            # Guardar o recuperar el Cliente principal
            cliente, creado = Cliente.objects.get_or_create(
                nit=nit,
                defaults={'nombre': nombre}
            )
            
            if creado:
                clientes_creados += 1
                self.stdout.write(f"Cliente creado: {nombre} (NIT: {nit})")
            else:
                cliente.nombre = nombre
                cliente.save()

            # 3. Buscar y enlazar las columnas de correos dinámicamente
            for col_name, col_idx in headers.items():
                if col_name.startswith('Correo') and col_idx < len(row):
                    correo_val = row[col_idx]
                    
                    if correo_val:
                        email_limpio = str(correo_val).strip()
                        
                        if email_limpio and email_limpio.lower() != 'none':
                            try:
                                # Validación estructural del correo electrónico
                                validate_email(email_limpio)
                                
                                # Guardar la relación en el modelo relacional ClienteCorreo
                                _, correo_creado = ClienteCorreo.objects.get_or_create(
                                    cliente=cliente,
                                    correo=email_limpio
                                )
                                if correo_creado:
                                    correos_registrados += 1
                                    
                            except ValidationError:
                                self.stdout.write(self.style.WARNING(
                                    f"Fila {row_idx}: Correo descartado por sintaxis inválida -> '{email_limpio}'"
                                ))

        # Cerramos el libro para liberar memoria
        wb.close()

        self.stdout.write(self.style.SUCCESS(
            f"\nProceso finalizado con éxito.\n"
            f"- Clientes procesados/nuevos: {clientes_creados}\n"
            f"- Direcciones de correo enlazadas: {correos_registrados}"
        ))